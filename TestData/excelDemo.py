import openpyxl

book = openpyxl.load_workbook("C:\\Users\\rauf.mirzoyev\\Desktop\\PythonDemo.xlsx") #otkril excel
sheet = book.active #aktiviroval sheet
Dict = {} #empty dictionary


cell = sheet.cell(row=1, column=2) #vibrali yacheyku
#print(cell.value)

sheet.cell(row=2, column=2).value = "rauf"


#print(sheet.cell(row=2, column=2).value)

#print(sheet.max_row)

#print(sheet.max_column)

#print(sheet['A5'].value)

for i in range(1, sheet.max_row+1): #to get rows
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, sheet.max_column+1):#to get columns
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            print("sadsadsad GIT2")
            print("sadsadsad GIT")
            print("sadsadsad GIT2")
            print("sadsadsad GIT")
            #Git demo project push request
            print("sadsadsad GIT2")
            print("sadsadsad GIT")
            print("sadsadsad GIT2")
            print("sadsadsad GIT")


print(Dict)

print("Branch test")


def gitDemoMergeConflict():
    pass

