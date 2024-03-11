import openpyxl


class HomePageData:
    test_HomePageData = [{"firstname": "Rauf", "email": "testste@asdkasdj.co", "password": "123456", "gender": "Male"},
                         {"firstname": "RaufTESTTEST", "email": "@asdkasdj.co", "password": "654123",
                          "gender": "Female"}]

    @staticmethod
    def get_test_data(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\rauf.mirzoyev\\Desktop\\PythonDemo.xlsx")  # otkril excel
        sheet = book.active  # aktiviroval sheet
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
